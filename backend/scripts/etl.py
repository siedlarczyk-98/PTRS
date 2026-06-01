"""
ETL: lê FPRS_Modelo3_sem_interacoes_v13.xlsx e popula o banco SQLite.
Uso: python scripts/etl.py <caminho_xlsx> [caminho_db]
"""
import sqlite3
import sys
from pathlib import Path
import openpyxl


def create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS medicamentos_base (
        nome_normalizado  TEXT PRIMARY KEY,
        classe_observacao TEXT,
        pim_beers         INTEGER NOT NULL DEFAULT 0,
        afinidade_ac      TEXT    NOT NULL DEFAULT 'No',
        afinidade_sedativa TEXT   NOT NULL DEFAULT 'No',
        peso_afinidade    INTEGER NOT NULL DEFAULT 0,
        fonte             TEXT
    );

    CREATE TABLE IF NOT EXISTS aliases (
        entrada_aceita   TEXT PRIMARY KEY,
        nome_normalizado TEXT NOT NULL REFERENCES medicamentos_base(nome_normalizado)
    );

    CREATE TABLE IF NOT EXISTS parametros (
        chave TEXT PRIMARY KEY,
        valor REAL NOT NULL,
        descricao TEXT
    );
    """)


def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _int(v) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _float(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


AFINIDADE_VALID = {"No", "Low", "Moderate", "High"}


def load_base_fixa(conn: sqlite3.Connection, ws) -> int:
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = _str(row[0]).lower()
        if not nome:
            continue
        classe = _str(row[1])
        pim = _int(row[2])
        afinidade_ac = _str(row[4]) if _str(row[4]) in AFINIDADE_VALID else "No"
        afinidade_sed = _str(row[6]) if _str(row[6]) in AFINIDADE_VALID else "No"
        peso = _int(row[7])
        fonte = _str(row[8])
        conn.execute(
            """INSERT OR REPLACE INTO medicamentos_base
               (nome_normalizado, classe_observacao, pim_beers, afinidade_ac, afinidade_sedativa, peso_afinidade, fonte)
               VALUES (?,?,?,?,?,?,?)""",
            (nome, classe, pim, afinidade_ac, afinidade_sed, peso, fonte),
        )
        count += 1
    return count


def load_aliases(conn: sqlite3.Connection, ws) -> int:
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        entrada = _str(row[0]).lower()
        normalizado = _str(row[1]).lower()
        if not entrada or not normalizado:
            continue
        conn.execute(
            "INSERT OR REPLACE INTO aliases (entrada_aceita, nome_normalizado) VALUES (?,?)",
            (entrada, normalizado),
        )
        count += 1
    return count


PARAMETROS_MAP = {
    # (texto parcial da linha, chave, descrição)
    "0-4": ("poli_0_4", "Polifarmácia 0–4 medicamentos"),
    "5-9": ("poli_5_9", "Polifarmácia 5–9 medicamentos"),
    "10": ("poli_10_mais", "Superpolifarmácia ≥10 medicamentos"),
    "baixa": ("afinidade_low", "Afinidade baixa"),
    "moderada": ("afinidade_moderate", "Afinidade moderada"),
    "alta": ("afinidade_high", "Afinidade alta"),
    "pim": ("pim_beers_adicional", "MPI/Beers adicional"),
    "capacidade": ("capacidade_max", "Capacidade máxima (medicamentos)"),
}


def load_parametros(conn: sqlite3.Connection, ws) -> int:
    # Insere os valores fixos/confirmados diretamente (fonte de verdade = especificação)
    params = [
        ("poli_0_4",            0.0,  "Polifarmácia 0–4 medicamentos"),
        ("poli_5_9",            0.5,  "Polifarmácia 5–9 medicamentos"),
        ("poli_10_mais",        1.0,  "Superpolifarmácia ≥10 medicamentos"),
        ("afinidade_low",       1.0,  "Afinidade baixa"),
        ("afinidade_moderate",  2.0,  "Afinidade moderada"),
        ("afinidade_high",      3.0,  "Afinidade alta"),
        ("pim_beers_adicional", 0.5,  "MPI/Beers adicional"),
        ("capacidade_max",      25.0, "Capacidade máxima (medicamentos)"),
        ("corte_alto_risco",    1.5,  "Ponto de corte alto risco"),
    ]
    for chave, valor, desc in params:
        conn.execute(
            "INSERT OR REPLACE INTO parametros (chave, valor, descricao) VALUES (?,?,?)",
            (chave, valor, desc),
        )
    return len(params)


def run_etl(xlsx_path: Path, db_path: Path) -> None:
    print(f"Lendo: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA foreign_keys = ON")

    create_schema(conn)

    n_base = load_base_fixa(conn, wb["Base_Fixa"])
    print(f"  medicamentos_base: {n_base} registros")

    n_alias = load_aliases(conn, wb["Alias_Map"])
    print(f"  aliases: {n_alias} registros")

    n_params = load_parametros(conn, wb["Parameters"])
    print(f"  parametros: {n_params} registros")

    conn.commit()
    conn.close()
    print(f"Banco gerado: {db_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/etl.py <caminho_xlsx> [caminho_db]")
        sys.exit(1)
    xlsx = Path(sys.argv[1])
    db = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("data/seed.db")
    db.parent.mkdir(parents=True, exist_ok=True)
    run_etl(xlsx, db)
